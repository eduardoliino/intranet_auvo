# app/colaborador_routes.py

from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file, jsonify, current_app
from flask_login import login_required
from app import db
from app.models import Colaborador, Cargo, Departamento
from app.admin_routes import admin_required, salvar_foto
from sqlalchemy.orm import joinedload
import pandas as pd
import io
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

colaborador_bp = Blueprint('colaborador', __name__,
                           url_prefix='/admin/colaboradores')


@colaborador_bp.route('/')
@admin_required
def listar():
    colaboradores = Colaborador.query.options(
        joinedload(Colaborador.cargo),
        joinedload(Colaborador.departamento)
    ).order_by(Colaborador.nome).all()

    colaboradores_json = [{
        'id': c.id, 'nome': c.nome, 'sobrenome': c.sobrenome,
        'email': c.email_corporativo,
        'cargo': c.cargo.titulo if c.cargo else '-',
        'cargo_id': c.cargo_id,
        'depto': c.departamento.nome if c.departamento else '-',
        'depto_id': c.departamento_id,
        'foto': c.foto_filename
    } for c in colaboradores]

    cargos = Cargo.query.order_by(Cargo.titulo).all()
    departamentos = Departamento.query.order_by(Departamento.nome).all()

    return render_template('admin/listar_colaboradores.html',
                           cargos=cargos,
                           departamentos=departamentos,
                           colaboradores_json=colaboradores_json)


@colaborador_bp.route('/adicionar', methods=['GET', 'POST'])
@admin_required
def adicionar():
    cargos = Cargo.query.order_by(Cargo.titulo).all()
    departamentos = Departamento.query.order_by(Departamento.nome).all()
    superiores = Colaborador.query.order_by(Colaborador.nome).all()

    if request.method == 'POST':
        form_data = request.form
        try:
            data_nascimento = pd.to_datetime(
                form_data.get('data_nascimento'), dayfirst=True).date()
            data_inicio = pd.to_datetime(form_data.get(
                'data_inicio'), dayfirst=True).date()
        except (ValueError, TypeError):
            flash('Data inválida. Por favor, verifique o dia, mês e ano.', 'danger')
            return render_template('admin/adicionar_colaborador_manual.html', form_data=form_data, cargos=cargos, departamentos=departamentos, superiores=superiores)

        email = form_data.get('email_corporativo')
        if Colaborador.query.filter_by(email_corporativo=email).first():
            flash(f'O e-mail "{email}" já está em uso.', 'danger')
            return render_template('admin/adicionar_colaborador_manual.html', form_data=form_data, cargos=cargos, departamentos=departamentos, superiores=superiores)

        foto_filename = None
        if 'foto' in request.files and request.files['foto'].filename != '':
            foto_filename = salvar_foto(request.files['foto'])

        novo_colaborador = Colaborador(
            nome=form_data.get('nome'),
            sobrenome=form_data.get('sobrenome'),
            email_corporativo=email,
            data_nascimento=data_nascimento,
            data_inicio=data_inicio,
            foto_filename=foto_filename,
            cargo_id=int(form_data.get('cargo_id')) if form_data.get(
                'cargo_id') else None,
            departamento_id=int(form_data.get('departamento_id')) if form_data.get(
                'departamento_id') else None,
            superior_id=int(form_data.get('superior_id')) if form_data.get(
                'superior_id') != 'None' else None
        )
        novo_colaborador.set_password(form_data.get('senha'))
        db.session.add(novo_colaborador)
        db.session.commit()
        flash('Colaborador adicionado com sucesso!', 'success')
        return redirect(url_for('colaborador.listar'))

    return render_template('admin/adicionar_colaborador_manual.html', cargos=cargos, departamentos=departamentos, superiores=superiores)


@colaborador_bp.route('/editar/<int:id>', methods=['GET', 'POST'])
@admin_required
def editar(id):
    colaborador = Colaborador.query.get_or_404(id)
    cargos = Cargo.query.order_by(Cargo.titulo).all()
    departamentos = Departamento.query.order_by(Departamento.nome).all()
    superiores = Colaborador.query.filter(
        Colaborador.id != id).order_by(Colaborador.nome).all()

    if request.method == 'POST':
        form_data = request.form
        try:
            data_nascimento = pd.to_datetime(
                form_data.get('data_nascimento'), dayfirst=True).date()
            data_inicio = pd.to_datetime(
                form_data.get('data_inicio'), dayfirst=True).date()
        except (ValueError, TypeError):
            flash('Data inválida. Por favor, verifique o dia, mês e ano.', 'danger')
            return render_template('admin/edit_colaborador.html', colaborador=colaborador, form_data=form_data, cargos=cargos, departamentos=departamentos, superiores=superiores)

        novo_superior_id = form_data.get('superior_id')
        novo_superior_id = int(
            novo_superior_id) if novo_superior_id and novo_superior_id != 'None' else None

        if is_circular_reference(id, novo_superior_id):
            flash(
                'Erro de hierarquia: Um colaborador não pode ser seu próprio superior.', 'danger')
            return render_template('admin/edit_colaborador.html', colaborador=colaborador, form_data=form_data, cargos=cargos, departamentos=departamentos, superiores=superiores)

        colaborador.nome = form_data.get('nome')
        colaborador.sobrenome = form_data.get('sobrenome')
        colaborador.email_corporativo = form_data.get('email_corporativo')
        colaborador.data_nascimento = data_nascimento
        colaborador.data_inicio = data_inicio
        colaborador.cargo_id = int(form_data.get(
            'cargo_id')) if form_data.get('cargo_id') else None
        colaborador.departamento_id = int(form_data.get(
            'departamento_id')) if form_data.get('departamento_id') else None
        colaborador.superior_id = novo_superior_id

        if 'foto' in request.files and request.files['foto'].filename != '':
            colaborador.foto_filename = salvar_foto(request.files['foto'])

        nova_senha = form_data.get('senha')
        if nova_senha:
            colaborador.set_password(nova_senha)

        db.session.commit()
        flash('Colaborador atualizado com sucesso!', 'success')
        return redirect(url_for('colaborador.listar'))

    return render_template('admin/edit_colaborador.html', colaborador=colaborador, cargos=cargos, departamentos=departamentos, superiores=superiores)

# As outras funções permanecem iguais


@colaborador_bp.route('/remover/<int:id>', methods=['POST'])
@admin_required
def remover(id):
    colaborador = Colaborador.query.get_or_404(id)
    if colaborador.foto_filename:
        try:
            caminho_foto = os.path.join(
                current_app.root_path, 'static/fotos_colaboradores', colaborador.foto_filename)
            if os.path.exists(caminho_foto):
                os.remove(caminho_foto)
        except Exception as e:
            print(f"Erro ao remover foto do colaborador {id}: {e}")

    nome_completo = f"{colaborador.nome} {colaborador.sobrenome}"
    db.session.delete(colaborador)
    db.session.commit()
    return jsonify({'success': True, 'message': f'Colaborador {nome_completo} e a sua foto foram removidos com sucesso!'})


@colaborador_bp.route('/importar', methods=['GET', 'POST'])
@admin_required
def importar():
    if request.method == 'POST':
        if 'planilha_colaboradores' not in request.files:
            flash('Nenhum ficheiro selecionado.', 'warning')
            return redirect(request.url)

        file = request.files['planilha_colaboradores']
        if file.filename == '':
            flash('Nenhum ficheiro selecionado.', 'warning')
            return redirect(request.url)

        if file and file.filename.endswith('.xlsx'):
            try:
                df = pd.read_excel(file, engine='openpyxl',
                                   header=0, skiprows=[1])
                novos_colaboradores = 0
                erros = []
                df.dropna(how='all', inplace=True)

                for index, row in df.iterrows():
                    try:
                        email = row.get('email_corporativo')
                        if pd.isna(email):
                            continue

                        if Colaborador.query.filter_by(email_corporativo=email).first():
                            erros.append(
                                f"Linha {index+3}: E-mail '{email}' já existe.")
                            continue

                        data_nascimento = pd.to_datetime(
                            row.get('data_nascimento'), dayfirst=True).date()
                        data_inicio = pd.to_datetime(
                            row.get('data_inicio'), dayfirst=True).date()

                        novo_colaborador = Colaborador(
                            nome=row.get('nome'),
                            sobrenome=row.get('sobrenome'),
                            email_corporativo=email,
                            data_nascimento=data_nascimento,
                            data_inicio=data_inicio
                        )
                        novo_colaborador.set_password(str(row.get('senha')))
                        db.session.add(novo_colaborador)
                        novos_colaboradores += 1
                    except (ValueError, TypeError):
                        erros.append(
                            f"Linha {index+3}: Data inválida. Verifique o formato e os valores.")
                        continue
                    except Exception as e:
                        erros.append(
                            f"Linha {index+3}: Erro nos dados. Detalhe: {e}")
                        continue

                db.session.commit()

                if novos_colaboradores > 0:
                    flash(
                        f'{novos_colaboradores} colaborador(es) importado(s) com sucesso!', 'success')
                if erros:
                    flash('Alguns registos não foram importados:', 'danger')
                    for erro in erros:
                        flash(erro, 'danger')

            except Exception as e:
                db.session.rollback()
                flash(
                    f'Ocorreu um erro ao processar o ficheiro: {e}', 'danger')

            return redirect(url_for('colaborador.listar'))

    return render_template('admin/importar_colaboradores.html')


@colaborador_bp.route('/baixar_modelo')
@admin_required
def baixar_modelo():
    dados_modelo = {
        'nome': ['João'],
        'sobrenome': ['Silva'],
        'email_corporativo': ['joao.silva@auvo.com'],
        'data_nascimento': ['15/10/1990'],
        'data_inicio': ['01/02/2024'],
        'senha': ['senha@123']
    }
    df_modelo = pd.DataFrame(dados_modelo)

    buffer = io.BytesIO()
    df_modelo.to_excel(buffer, index=False,
                       sheet_name='colaboradores', engine='openpyxl')
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb.active

    header_fill = PatternFill(start_color="3A1B4A",
                              end_color="3A1B4A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    example_fill = PatternFill(
        start_color="E8CCF6", end_color="E8CCF6", fill_type="solid")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for cell in ws[2]:
        cell.fill = example_fill

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 5)
        ws.column_dimensions[column].width = adjusted_width

    final_buffer = io.BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)

    return send_file(
        final_buffer,
        as_attachment=True,
        download_name='modelo_importacao_colaboradores.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def is_circular_reference(colaborador_id, novo_superior_id):
    if not novo_superior_id or not colaborador_id:
        return False
    if int(colaborador_id) == int(novo_superior_id):
        return True
    visitados = set()
    atual = Colaborador.query.get(novo_superior_id)
    while atual:
        if atual.id in visitados:
            break
        visitados.add(atual.id)
        if atual.superior_id == int(colaborador_id):
            return True
        atual = atual.superior
    return False
